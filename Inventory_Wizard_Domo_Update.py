import csv
from openpyxl import Workbook
import openpyxl
import datetime
import re
from openpyxl.styles.numbers import FORMAT_PERCENTAGE

def main():
    Wizard = []
    Inventory = []
    Item_Information = []
    Shipments = []

    # Read Inventory Wizard File From Domo
    with open('Inventory_Wizard.csv','r',encoding='utf8') as wizard:
        reader = csv.DictReader(wizard)
        for row in reader:
            Wizard.append({
                'Customer':row['Customer'],'P.O. Number':row['P.O. Number'],'NAI P.O.':row['NAI P.O.'],'Container #':row['Container #'],
                'Document Number': row['Document Number'], 'Sell Price': row['Sell Price'], 'Item':row['Item'],
                'Location':row['Location'], 'Schedule Date' : row['Schedule Date'],'Quantity':row['Quantity'],
                'Actual Inventory Level': row['Actual Inventory Level'], 'Type': row['Type'], 'Status': row['Status'],'Initial Forecast':row['Initial Forecast'],
                'Forecast Remaining' : row['Forecast Remaining'],'Forecast':row['Forecast'], 'Forecast Inventory Level': row['Forecast Inventory Level'],'Sort':row['Sort'],
                })

    #Read Inventory in appendix
    with open('Inventory.csv','r') as inventory:
        reader = csv.DictReader(inventory)
        for row in reader: 
            Inventory.append({
                'Item': row['Item'],'Location':row['Location'],'On Hand': row['On Hand'] ,
            })
    
    #Read Item_information
    with open('Item_Information.csv','r') as item_information:
        reader = csv.DictReader(item_information)
        for row in reader: 
            Item_Information.append({
                'Unique ID': row['Unique ID'], 'Display Name': row['Display Name'],'Product Group': row['Product Group'], 'Program Year': row['Program Year'], 
                'Item Status': row['Item Status'], 'Master Pack Quantity' : row['Master Pack Quantity'], 'Exclusivity': row['Exclusivity'],'Factory':row['Factory'],
                'Old SKU' : row['Old SKU'], 'New SKU' : row['New SKU'],'Last Purchase Price':row['Last Purchase Price'] , 'Purchase Price': row['Purchase Price'],
                'Unclaimed Inventory' : row['Unclaimed Inventory'], "QTY PER 40'HQ CONTAINER" : row["QTY PER 40'HQ CONTAINER"],
            })

    #Read Shipment information
    with open('Shipments.csv','r') as shipments:
        reader = csv.DictReader(shipments)
        for row in reader: 
            Shipments.append({
                'Item':row['Item'], 'Import/Domestic': row['Import/Domestic'],'LY Quantity': row['LY Quantity'], 'YTD Quantity': row['YTD Quantity'], 
                'MTD Quantity': row['MTD Quantity'], 'LY Cancelled' : row['LY Cancelled'], 'YTD Cancelled' : row['YTD Cancelled'],'MTD Cancelled' : row['MTD Cancelled']
                , 'LY YTD Quantity' : row['LY YTD Quantity']
            })


    #create file name
    book_name = 'NAI_Inventory_Wizard' + ' ' + str(datetime.datetime.today().strftime("%Y-%m-%d %A %H;%M;%S"))+".xlsx"
    #date = datetime.date.isoformat(datetime.date.today())
    #time = datetime.datetime.today().__format__('%H:%M:%S')

    #create workbook and unique SKU list as well as sort
    wb = Workbook()
    unique_list = []

    for item in sorted(Wizard, key = lambda Wizard : Wizard['Item']):
        if item['Customer'] == 'Beginning Inventory' and int(float(item['Quantity'])) == 0:
            pass
        elif item['Item'] not in unique_list:
            unique_list.append(item['Item'])

    #create common formatting techniques for use throughout
    
    thin_border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
    #thick_border = openpyxl.styles.borders.Border(
                #left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thick'),
                #top=openpyxl.styles.borders.Side(style='thick')
                #)
    center = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold = openpyxl.styles.Font(bold=True)
    underline = openpyxl.styles.Font(underline = 'single')
    double_underline = openpyxl.styles.Font(underline = 'double')
    right = openpyxl.styles.Alignment(horizontal='right', vertical='center', wrap_text=False)

    

    #Read header names
    all_keys = Wizard[0].keys()
    all_keys = list(all_keys)
    all_keys.remove(all_keys[-1])

    #start of worksheet 
    for SKU in unique_list:
        ws = wb.create_sheet(title=SKU)
        count = 1 
        #header rows created here 
        header_row = 13
        for i in all_keys:
            ws.cell(row=header_row, column = count, value = i)
            ws.cell(row=header_row, column=count).alignment = center
            ws.cell(row=header_row, column=count).font = bold
            ws.cell(row=header_row, column=count).border = thin_border
            count += 1

        #column widths set here
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 11
        ws.column_dimensions['F'].width = 6
        ws.column_dimensions['G'].width = 7
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 11
        ws.column_dimensions['J'].width = 9
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 25
        ws.column_dimensions['M'].width = 18
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 11
        ws.column_dimensions['P'].width = 11
        ws.column_dimensions['Q'].width = 11
        ws.column_dimensions['R'].width = 28
        ws.column_dimensions['S'].width = 15

        ws.freeze_panes = 'A14'

        current_row = header_row + 1
        #Sorting
        sort = sorted(Wizard, key = lambda Wizard:Wizard['NAI P.O.'])
        sort = sorted(sort, key = lambda sort:sort['P.O. Number'])
        sort = sorted(sort,key = lambda sort:sort['Schedule Date'])
        sort = sorted(sort,key = lambda sort:sort['Sort'],reverse = True)
        #Start of Item Specifics 
        for row in sort: 
            if row['Item'] == SKU:
                #convert Text to Numbers
                try:
                    sell_price = float(row['Sell Price'])
                except ValueError:
                    sell_price = row['Sell Price']
                try:
                    Item = int(float(row['Item']))
                except ValueError:
                    Item = row['Item']
                try:
                    Quantity = int(float(row['Quantity']))
                except ValueError:
                    Quantity = row['Quantity']
                try:
                    Actual_Inventory_Level = int(float(row['Actual Inventory Level']))
                except ValueError:
                    Actual_Inventory_Level = row['Actual Inventory Level']
                try:
                    Initial_Forecast = int(float(row['Initial Forecast']))
                except ValueError:
                    Initial_Forecast = row['Initial Forecast']
                try:
                    Forecast_Remaining = int(float(row['Forecast Remaining']))
                except ValueError:
                    Forecast_Remaining = row['Forecast Remaining']
                try:
                    Forecast = int(float(row['Forecast']))
                except ValueError:
                    Forecast = row['Forecast']
                try:
                    if row['Type'] == 'Beginning Inventory': 
                        Forecast_Inventory_Level = int(float(row['Forecast Inventory Level']))
                    else: 
                        Forecast_Inventory_Level = '=IF(L'+str(current_row)+'="P.O.",Q'+str(current_row-1)+',IF(H'+str(current_row)+'="Import",Q'+str(current_row-1)+',IF(L'+str(current_row)+'="Forecast",Q'+str(current_row-1)+'+(P'+str(current_row)+'*$Q$12),Q'+str(current_row-1)+'+P'+str(current_row)+')))'
                        
                except ValueError:
                    Forecast_Inventory_Level = row['Forecast Inventory Level']
                
                #Asign Wizard Numbers to cells
                ws.cell(row=current_row, column=1, value=row['Customer'])
                ws.cell(row=current_row, column=2, value=row['P.O. Number'])
                ws.cell(row=current_row, column=3, value=row['NAI P.O.'])
                ws.cell(row=current_row, column=4, value=row['Container #'])
                ws.cell(row=current_row, column=5, value=row['Document Number'])
                ws.cell(row=current_row, column=6, value= sell_price)
                ws.cell(row=current_row, column=7, value= Item)
                ws.cell(row=current_row, column=8, value=row['Location'])
                ws.cell(row=current_row, column=9, value=row['Schedule Date'])
                ws.cell(row=current_row, column=10, value=Quantity)
                ws.cell(row=current_row, column=11, value=Actual_Inventory_Level)
                ws.cell(row=current_row, column=12, value=row['Type'])
                ws.cell(row=current_row, column=13, value=row['Status'])
                ws.cell(row=current_row, column=14, value=Initial_Forecast)
                ws.cell(row=current_row, column=15, value=Forecast_Remaining)
                ws.cell(row=current_row, column=16, value=Forecast)
                ws.cell(row=current_row, column=17, value=Forecast_Inventory_Level)
                
                current_row += 1
            else:
                pass

        #Sum Inventory positions
        BPUT1 = 0
        BPUT2 = 0
        BPUT3 = 0
        All_Roads_Trucking = 0
        Keyword_SLC = 0 
        Clearfield = 0 
        Partners_Trade = 0
        Damaged_Location = 0 
        Rework = 0 

        for row in Inventory:
            if row['Item'] == SKU:
                if row['Location'] == 'BPUT1':
                    BPUT1 += int(float(row['On Hand']))
                elif row['Location'] == 'BPUT2':
                    BPUT2 += int(float(row['On Hand']))
                elif row['Location'] == 'BPUT3':
                    BPUT3 += int(float(row['On Hand']))
                elif row['Location'] == 'All Roads Trucking':
                    All_Roads_Trucking += int(float(row['On Hand']))
                elif row['Location'] == 'Keyword SLC':
                    Keyword_SLC += int(float(row['On Hand']))
                elif row['Location'] == 'Clearfield':
                    Clearfield += int(float(row['On Hand']))
                elif row['Location'] == 'Partners Trade':
                    Partners_Trade += int(float(row['On Hand']))
                elif row['Location'] == 'Damaged Location BP':
                    Damaged_Location += int(float(row['On Hand']))
                elif row['Location'] == 'Rework':
                    Rework += int(float(row['On Hand']))



        


        #Assign Inventory Values

        ws.cell(row=4,column=7,value='Current Instock by Location:')
        ws.cell(row=4,column=7).font = bold
        ws.cell(row=4,column=7).alignment = right 
        
        ws.cell(row=5,column=6,value='BPUT1:')
        ws.cell(row=5,column=6).font = underline
        ws.cell(row=5,column=6).alignment = right
        ws.cell(row=5,column=7).alignment = center
        ws.cell(row=5,column=7,value=BPUT1)

        ws.cell(row=6,column=6,value='BPUT2:')
        ws.cell(row=6,column=6).font = underline
        ws.cell(row=6,column=6).alignment = right
        ws.cell(row=6,column=7).alignment = center
        ws.cell(row=6,column=7,value=BPUT2)

        ws.cell(row=7,column=6,value='BPUT3:')
        ws.cell(row=7,column=6).font = underline
        ws.cell(row=7,column=6).alignment = right
        ws.cell(row=7,column=7).alignment = center
        ws.cell(row=7,column=7,value=BPUT3)

        ws.cell(row=5,column=9,value='All Roads Trucking:')
        ws.cell(row=5,column=9).font = underline
        ws.cell(row=5,column=9).alignment = right
        ws.cell(row=5,column=10).alignment = center
        ws.cell(row=5,column=10,value=All_Roads_Trucking)

        ws.cell(row=6,column=9,value='Keyword SLC:')
        ws.cell(row=6,column=9).font = underline
        ws.cell(row=6,column=9).alignment = right
        ws.cell(row=6,column=10).alignment = center
        ws.cell(row=6,column=10,value=Keyword_SLC)

        ws.cell(row=7,column=9,value='Clearfield:')
        ws.cell(row=7,column=9).font = underline
        ws.cell(row=7,column=9).alignment = right
        ws.cell(row=7,column=10).alignment = center
        ws.cell(row=7,column=10,value=Clearfield)

        ws.cell(row=5,column=12,value='Partners Trade:')
        ws.cell(row=5,column=12).font = underline
        ws.cell(row=5,column=12).alignment = right
        ws.cell(row=5,column=13).alignment = center
        ws.cell(row=5,column=13,value=Partners_Trade)

        ws.cell(row=6,column=12,value='Damaged Location:')
        ws.cell(row=6,column=12).font = underline
        ws.cell(row=6,column=12).alignment = right
        ws.cell(row=6,column=13).alignment = center
        ws.cell(row=6,column=13,value=Damaged_Location)

        ws.cell(row=7,column=12,value='Rework:')
        ws.cell(row=7,column=12).font = underline
        ws.cell(row=7,column=12).alignment = right
        ws.cell(row=7,column=13).alignment = center
        ws.cell(row=7,column=13,value=Rework)
        



        #Asign Item Information to cells
        ws.cell(row=1,column=1,value='Item Number:') 
        ws.cell(row=1,column=1).alignment = center

        ws.cell(row=1,column=2,value=SKU)
        ws.cell(row=1,column=2).alignment = center
        ws.cell(row=1,column=2).font = bold

        #Item Information Labels       
        ws.cell(row=4,column=1,value='Product Group:')
        ws.cell(row=5,column=1,value='Program Year:')
        ws.cell(row=6,column=1,value='Item Status:')
        ws.cell(row=7,column=1,value='Master Pack Quantity:')
        ws.cell(row=8,column=1,value='Exclusivity:')
        ws.cell(row=9,column=1,value='Factory:')


        Purchase_Price = ""
        Last_Purchase_Price = ""
        unclaimed_inventory = 0 
        Container_Qty = ''

        #Item Information Values
        for item in Item_Information:
            if item['Unique ID'] == SKU:
                try:
                    MP_Quantity = str(int(float(item['Master Pack Quantity'])))
                except:
                    MP_Quantity = item['Master Pack Quantity']
                try:
                    Purchase_Price = float(item['Purchase Price'])
                except ValueError:
                    pass
                try:
                    Last_Purchase_Price = float(item['Last Purchase Price'])
                except ValueError:
                    pass
                try:
                    unclaimed_inventory = float(item['Unclaimed Inventory'])
                except ValueError:
                    pass

                ws.cell(row=2,column=1,value=item['Display Name'])
                ws.cell(row=4,column=2,value=item['Product Group'])
                ws.cell(row=5,column=2,value=item['Program Year'])
                ws.cell(row=6,column=2,value=item['Item Status'])
                ws.cell(row=7,column=2,value= MP_Quantity)
                ws.cell(row=8,column=2,value=item['Exclusivity'])
                ws.cell(row=9,column=2,value=item['Factory'])
                ws.cell(row=2,column=1).font = bold
                ws.cell(row=4,column=2).font = bold
                ws.cell(row=5,column=2).font = bold
                ws.cell(row=6,column=2).font = bold
                ws.cell(row=7,column=2).font = bold
                ws.cell(row=8,column=2).font = bold
                ws.cell(row=9,column=2).font = bold

                #Unclaimed Inventory
                ws.cell(row=9,column=12,value='Factory Inventory:').font = underline
                ws.cell(row=9,column=12).alignment = right
                ws.cell(row=9,column=13,value = unclaimed_inventory).alignment = center

                #40GP Conatiner Qty
                ws.cell(row=10,column=1,value="40 HQ Container Qty")
                ws.cell(row=10,column=2,value=item["QTY PER 40'HQ CONTAINER"]).font = bold


                #Old SKU, New SKU
                ws.cell(row=1,column=4,value='Old SKU(s):')
                ws.cell(row=2,column=4,value='New SKU(s):')

                try: 
                    old_skus = re.sub(',','',item['Old SKU'])
                except: 
                    old_skus = []

                count = 0
                old_skus2 = []
                old_sku = ''
                for number in old_skus:
                    if count == 3:
                        old_sku += number
                        count = 0
                        old_skus2.append(old_sku)
                        old_sku = ''
                    else:
                        old_sku += number
                        count += 1

                column = 5

                for sku in old_skus2:
                    ws.cell(row=1,column=column,value=sku).hyperlink = (book_name + '#' + sku + '!' + 'A1')
                    column += 1

                try: 
                    new_skus = re.sub(',','',item['New SKU'])
                except: 
                    new_skus = []

                count = 0
                new_skus2 = []
                new_sku = ''
                for number in new_skus:
                    if count == 3:
                        new_sku += number
                        count = 0
                        new_skus2.append(new_sku)
                        new_sku = ''
                    else:
                        new_sku += number
                        count += 1

                column = 5
                
                for sku in new_skus2:
                    ws.cell(row=2,column=column,value=sku).hyperlink = (book_name + '#' + sku + '!' + 'A1')
                    column += 1


        #Asign Purchase Price and Last Purchase Price to cells 
        ws.cell(row=10,column=4,value='Purchase Price').alignment = right
        ws.cell(row=11,column=4,value='Last Purchase Price').alignment = right

        ws.cell(row=10,column=5,value=Purchase_Price)
        ws.cell(row=11,column=5,value=Last_Purchase_Price)

        #Shipment Labels
        ws.cell(row=3,column=15,value='Import')
        ws.cell(row=3,column=15).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=False)
        ws.cell(row=3,column=15).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=3,column=16,value='Domestic')
        ws.cell(row=3,column=16).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=False)
        ws.cell(row=3,column=16).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=3,column=17,value='Total')
        ws.cell(row=3,column=17).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=False)
        ws.cell(row=3,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=4,column=14,value='LY Shipped')
        ws.cell(row=4,column=14).alignment = center
        #ws.cell(row=9,column=13).font = openpyxl.styles.Font(size=8)
        ws.cell(row=4,column=14).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=5,column=14,value='LY YTD Shipped')
        ws.cell(row=5,column=14).alignment = center
        #ws.cell(row=10,column=13).font = openpyxl.styles.Font(size=8)
        ws.cell(row=5,column=14).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=6,column=14,value='YTD Shipped')
        ws.cell(row=6,column=14).alignment = center
        #ws.cell(row=10,column=13).font = openpyxl.styles.Font(size=8)
        ws.cell(row=6,column=14).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=7,column=14,value='MTD Shipped')
        ws.cell(row=7,column=14).alignment = center
        #ws.cell(row=11,column=13).font = openpyxl.styles.Font(size=8)
        ws.cell(row=7,column=14).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=8,column=14,value='LY Cancelled').border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=8,column=14).alignment = center
        ws.cell(row=9,column=14,value='YTD Cancelled').border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=9,column=14).alignment = center
        ws.cell(row=10,column=14,value='MTD Cancelled').border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=10,column=14).alignment = center


        #Assign Shipment Values
        import_LY = 0
        import_YTD = 0 
        import_MTD = 0 
        domestic_LY = 0 
        domestic_YTD = 0 
        domestic_MTD = 0 
        total_LY = 0 
        total_YTD = 0 
        total_MTD = 0 
        import_LY_C = 0
        import_YTD_C = 0 
        import_MTD_C = 0 
        domestic_LY_C = 0
        domestic_YTD_C = 0
        domestic_MTD_C = 0 
        total_LY_C = 0 
        total_YTD_C = 0 
        total_MTD_C = 0 
        import_ly_ytd = 0
        domestic_ly_ytd = 0 
        total_ly_ytd = 0 


        for shipment in Shipments: 
            if shipment['Item'] == SKU:
                if shipment['Import/Domestic'] == 'Import':
                    try: 
                        import_LY += int(float(shipment['LY Quantity']))
                    except ValueError:
                        pass
                    try: 
                        import_YTD += int(float(shipment['YTD Quantity']))
                    except ValueError:
                        pass
                    try: 
                        import_MTD += int(float(shipment['MTD Quantity']))
                    except ValueError:
                        pass
                    try:
                        import_LY_C += int(float(shipment['LY Cancelled']))
                    except ValueError:
                        pass
                    try:
                        import_YTD_C += int(float(shipment['YTD Cancelled']))
                    except ValueError:
                        pass
                    try:
                        import_MTD_C += int(float(shipment['MTD Cancelled']))
                    except ValueError:
                        pass
                    try:
                        import_ly_ytd += int(float(shipment['LY YTD Quantity']))
                    except ValueError:
                        pass
                else:
                    try: 
                        domestic_LY += int(float(shipment['LY Quantity']))
                    except ValueError:
                        pass
                    try:
                        domestic_YTD += int(float(shipment['YTD Quantity']))
                    except ValueError:
                        pass
                    try:
                        domestic_MTD += int(float(shipment['MTD Quantity']))
                    except ValueError:
                        pass
                    try:
                        domestic_LY_C += int(float(shipment['LY Cancelled']))
                    except: 
                        pass
                    try:
                        domestic_YTD_C += int(float(shipment['YTD Cancelled']))
                    except:
                        pass
                    try:
                        domestic_MTD_C += int(float(shipment['MTD Cancelled']))
                    except:
                        pass
                    try:
                        domestic_ly_ytd += int(float(shipment['LY YTD Quantity']))
                    except:
                        pass
                
                total_LY = import_LY + domestic_LY
                total_YTD = import_YTD + domestic_YTD
                total_MTD = import_MTD + domestic_MTD
                total_LY_C = import_LY_C + domestic_LY_C
                total_YTD_C = import_YTD_C + domestic_YTD_C
                total_MTD_C = import_MTD_C + domestic_MTD_C
                total_ly_ytd = import_ly_ytd + domestic_ly_ytd

        #Asign Shipment Values to Cells
        ws.cell(row=4,column=15,value=import_LY).alignment = center
        ws.cell(row=5,column=15,value=import_ly_ytd).alignment = center
        ws.cell(row=6,column=15,value=import_YTD).alignment = center
        ws.cell(row=7,column=15,value=import_MTD).alignment = center  
        ws.cell(row=4,column=16,value=domestic_LY).alignment = center
        ws.cell(row=5,column=16,value=domestic_ly_ytd).alignment = center  
        ws.cell(row=6,column=16,value=domestic_YTD).alignment = center    
        ws.cell(row=7,column=16,value=domestic_MTD).alignment = center 
        ws.cell(row=4,column=17,value=total_LY).alignment = center
        ws.cell(row=5,column=17,value=total_ly_ytd).alignment = center   
        ws.cell(row=6,column=17,value=total_YTD).alignment = center    
        ws.cell(row=7,column=17,value=total_MTD).alignment = center

        ws.cell(row=8,column=15,value=import_LY_C).alignment = center
        ws.cell(row=8,column=16,value=domestic_LY_C).alignment = center  
        ws.cell(row=8,column=17,value=total_LY_C).alignment = center
        ws.cell(row=8,column=15).border = thin_border
        ws.cell(row=8,column=16).border = thin_border
        ws.cell(row=8,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )

        ws.cell(row=9,column=15,value=import_YTD_C).alignment = center
        ws.cell(row=9,column=16,value=domestic_YTD_C).alignment = center  
        ws.cell(row=9,column=17,value=total_YTD_C).alignment = center
        ws.cell(row=9,column=15).border = thin_border
        ws.cell(row=9,column=16).border = thin_border
        ws.cell(row=9,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )

        ws.cell(row=10,column=15,value=import_MTD_C).alignment = center
        ws.cell(row=10,column=16,value=domestic_MTD_C).alignment = center  
        ws.cell(row=10,column=17,value=total_MTD_C).alignment = center
        ws.cell(row=10,column=15).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=10,column=16).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=10,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )

        ws.cell(row=4,column=15).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=5,column=15).border = thin_border
        ws.cell(row=6,column=15).border = thin_border
        ws.cell(row=7,column=15).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=4,column=16).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=5,column=16).border = thin_border
        ws.cell(row=6,column=16).border = thin_border
        ws.cell(row=7,column=16).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=4,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=5,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=6,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=7,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )

        #Asign Summary Numbers 
        onhand = 0
        domestic_SO = 0 
        import_SO = 0 
        domestic_PO = 0
        import_PO = 0
        transit = 0 
        available = 0
        import_forecast = 0 
        domestic_forecast = 0

        for row in Wizard:
            if SKU == row['Item']:
                if row['Type'] == 'Beginning Inventory':
                    onhand = int(float(row['Quantity']))

                elif row['Type'].__contains__('S.O.'):
                    if row['Location'] == 'Import':
                        import_SO += int(float(row['Quantity']))*-1
                    else:
                        domestic_SO += int(float(row['Quantity']))*-1
                
                elif row['Type'] == 'P.O.' or row['Type'] == 'Non-Confirmed Blanket P.O.':
                    import_PO += int(float(row['Quantity']))
                
                elif row['Type'] == 'Confirmed P.O.' or row['Type'] == 'Non-Confirmed P.O.':
                    domestic_PO += int(float(row['Quantity']))

                elif row['Type'] == 'Est. Transit Recieve':
                    transit += int(float(row['Quantity']))
                
                elif row['Location'] == 'Import': 
                    import_forecast += int(float(row['Forecast']))*-1
                
                else:
                    domestic_forecast += int(float(row['Forecast']))*-1

        available = onhand - domestic_SO
        if available < 0: 
            available = 0

        #Asign Summary Numbers to Cells
        ws.cell(row=3,column=18,value='Total Onhand:').font = double_underline
        ws.cell(row=4,column=18,value='Domestic Sales on Order:').font = double_underline
        ws.cell(row=5,column=18,value='Import Sales on Order:').font = double_underline
        ws.cell(row=6,column=18,value='Domestic Purchase on Order:').font = double_underline
        ws.cell(row=7,column=18,value='Import Purchase on Order:').font = double_underline
        ws.cell(row=8,column=18,value='Transit:').font = double_underline
        ws.cell(row=9,column=18,value='Available:').font = double_underline
        ws.cell(row=10,column=18,value='Domestic Forecast:').font = double_underline
        ws.cell(row=11,column=18,value='Import Forecast:').font = double_underline

        ws.cell(row=3,column=18).alignment = right
        ws.cell(row=4,column=18).alignment = right
        ws.cell(row=5,column=18).alignment = right
        ws.cell(row=6,column=18).alignment = right
        ws.cell(row=7,column=18).alignment = right
        ws.cell(row=8,column=18).alignment = right
        ws.cell(row=9,column=18).alignment = right
        ws.cell(row=10,column=18).alignment = right
        ws.cell(row=11,column=18).alignment = right

        ws.cell(row=3,column=19,value=onhand).alignment = center
        ws.cell(row=4,column=19,value=domestic_SO).alignment = center
        ws.cell(row=5,column=19,value=import_SO).alignment = center
        ws.cell(row=6,column=19,value=domestic_PO).alignment = center
        ws.cell(row=7,column=19,value=import_PO).alignment = center
        ws.cell(row=8,column=19,value=transit).alignment = center
        ws.cell(row=9,column=19,value=available).alignment = center
        ws.cell(row=10,column=19,value=domestic_forecast).alignment = center
        ws.cell(row=11,column=19,value=import_forecast).alignment = center



        #Stock below 0 date Calculations
        actual_below0 = ''
        actual_days = ''
        forecast_below0 = ''
        forecast_days = ''
        for row in Wizard:
            if SKU == row['Item']: 
                if float(row['Actual Inventory Level']) < 0: 
                    actual_below0 = row['Schedule Date']
                    actual_belowedit = re.split('-',actual_below0)
                    actual_days = int((datetime.date(int(actual_belowedit[0]),int(actual_belowedit[1]),int(actual_belowedit[2])) - datetime.date.today()).days)
                    break

        for row in Wizard:
            if SKU == row['Item']:
                if float(row['Forecast Inventory Level']) < 0:
                    forecast_below0 = row['Schedule Date']
                    forecast_belowedit = re.split('-',forecast_below0)
                    forecast_days = int((datetime.date(int(forecast_belowedit[0]),int(forecast_belowedit[1]),int(forecast_belowedit[2])) - datetime.date.today()).days)
                    break
        
        #ws.cell(row=3,column=11,value='Run Date:')
        #ws.cell(row=4,column=11,value='Run Time:')
        #ws.cell(row=3,column=12,value=date)
        #ws.cell(row=4,column=12,value=time)

        ws.cell(row=10,column=7,value='Actual').border = thin_border
        ws.cell(row=11,column=7,value='Forecast').border = thin_border
        ws.cell(row=9,column=8,value='Stock below 0 Date').border = thin_border
        ws.cell(row=9,column=9,value='Days').border = thin_border

        ws.cell(row=10,column=8,value=actual_below0).border = thin_border
        ws.cell(row=11,column=8,value=forecast_below0).border = thin_border
        ws.cell(row=10,column=9,value=actual_days).border = thin_border
        ws.cell(row=11,column=9,value=forecast_days).border = thin_border
        ws.cell(row=11, column=7).font = openpyxl.styles.Font(size=9)
        ws.cell(row=10, column=7).font = openpyxl.styles.Font(size=9)

        # Demand Multiplier
        ws.cell(row=12,column=17,value=1).number_format = FORMAT_PERCENTAGE
        ws.cell(row=12,column=15,value='Demand Multiplier')
        ws.merge_cells(start_row=12,start_column=15,end_row=12,end_column=16)

        
    #save workbook
    wb.save(book_name)

     
        
    
if __name__ == '__main__':
    main()
