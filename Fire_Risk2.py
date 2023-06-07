from pdfquery import PDFQuery
import os
from openpyxl import Workbook
from datetime import datetime

def extract_annotations(pdf):
    annotations = []
    for annot in pdf.pq('Annot'):
        annot_attributes = annot.attrib
        annotations.append(annot_attributes)
    return annotations

def extract_pdf_annotations(pdf_file_path):
    with open(pdf_file_path, "rb") as file:
        pdf = PDFQuery(file)
        pdf.load()
        annotations = extract_annotations(pdf)
        return annotations

### Change your folder path ####
directory = r"C:\Users\Keaton\Documents\Fire Risk Management"

current_row = 2 
current_time = datetime.now().strftime("%Y-%m-%d%H-%M-%S")
wb = Workbook()
ws = wb.active
ws['A1'] = 'Incident Date & Time'
ws['B1'] = 'Case Number'
ws['C1'] = 'Griddle Model Number'
ws['D1'] = 'Griddle Serial Number'
for filename in os.listdir(directory):
    if filename.endswith(".pdf"):
        annotations = extract_pdf_annotations(filename)
        x = {}
        y = ['Incident Date','Case #','Model #','Serial #']
            

        for i in y: 
            for a in range(len(annotations)):
                try:
                    if annotations[a]['T'] == i:
                        x[i] = annotations[a]['V']
                except:
                    pass
                


        try:
            ws['A'+ str(current_row)] = x['Incident Date']
        except:
            pass
        try:
            ws['B'+ str(current_row)] = x['Case #']
        except:
            pass
        try:
            ws['C'+ str(current_row)] = x['Model #']
        except:
            pass
        try:
            ws['D'+ str(current_row)] = x['Serial #']
        except: pass

        current_row += 1
        if x == {}:
            print(filename,'is not in correct format')
        
        
        
wb.save('Fire_Risk_compiled_on_' + current_time + '.xlsx')
        

